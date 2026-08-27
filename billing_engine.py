"""
billing_engine.py
==================
Web-native re-implementation of the calculation/write core of
orderSummary.py + mtsBills.py, for the dashboard's "Bills" page.

Faithfully ported (verified line-by-line against orderSummary.py):
    - rateWithout_gst()     : taxable rate/unit = price * 100 / (100 + gst%)
    - tax bucketing         : lines grouped by integer GST%, CGST == SGST
                               always (half-tax rounded ROUND_HALF_UP), see
                               get_tax_calculation() / linewise_tax_bucket()
    - bottom_part()          : ttaxamt/cgst/sgst/total, "total = ttaxamt"
                               only for non-"Tax Invoice" bill types
    - margin                 : qty * (rate - cost_price) per line, matching
                               record_room_lookup.compute_margin_for_bill_json
    - Item_Description build : "{sheet} {Item_Name} {Item_Description}",
                               whitespace-squeezed
    - HSN truncation          : hsn[:4]
    - amount-in-words         : num2words(lang='en_IN')
    - sales_log.xlsx columns  : A fileName / B billType / C Date / D Invoice /
                               E customerID / F customerName / G total /
                               H margin / I status / J amount-breakup /
                               K taxable total / L taxable-breakup /
                               M margin-breakup / O Candidate

Deliberately NOT ported (out of scope until real assets are provided):
    - .docx template rendering (needs the real templateBook/*.docx files)
    - the CurrentQTY-scratchpad desktop workflow (superseded by this web form)
    - indexBook/invoiceNum.txt (falls back to scanning sales_log.xlsx for the
      last invoice number and incrementing its last 4 digits, exactly the
      way get_indexNum() does internally - drop the real invoiceNum.txt into
      indexBook/ and this module can be pointed at it for exact parity)
"""
import copy
import json
import os
import re
from decimal import Decimal, ROUND_HALF_UP

import openpyxl
from num2words import num2words

import record_room_lookup as rrl

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
RECORD_ROOM_BASE = BASE_DIR
INDEX_BOOK_DIR = os.path.join(BASE_DIR, "indexBook")

PRODUCT_SHEET_SKIP = {"HOME", "CUSTPR", "VENDOR"}

PRODUCT_COLUMNS = [
    "Item_Code", "Item_Description", "HSN_Code", "Item_Name", "CurrentQTY",
    "MPR", "Quantity", "GST %", "Rate", "Cost_Price", "Sale_RateT1", "Sale_PriceT1",
]

CUSTOMER_COLUMNS = ["customerID", "Customer_Name", "Contact_Number", "Address_Details", "GSTN"]


# --------------------------------------------------------------------------- #
# Catalog: products + customers, read straight from billdesk.xlsx
# --------------------------------------------------------------------------- #
def load_product_catalog(billdesk_path):
    """
    Returns a flat list of every product row across every non-special sheet:
        {sheet, row, item_code, item_name, item_description, hsn, mrp,
         quantity (live stock), gst_pct, cost_price, sale_price}
    `row` is the 1-based row number in that sheet - needed to write the
    stock decrement back after a sale.
    """
    wb = openpyxl.load_workbook(billdesk_path, data_only=True)
    catalog = []
    for sheet_name in wb.sheetnames:
        if sheet_name in PRODUCT_SHEET_SKIP:
            continue
        sht = wb[sheet_name]
        headers = {}
        for j in range(1, sht.max_column + 1):
            h = sht.cell(row=1, column=j).value
            if h in PRODUCT_COLUMNS:
                headers[h] = j
        if "Item_Name" not in headers:
            continue

        for i in range(2, sht.max_row + 1):
            item_name = sht.cell(row=i, column=headers["Item_Name"]).value
            if item_name in (None, ""):
                continue

            def cell(col_name):
                col = headers.get(col_name)
                return sht.cell(row=i, column=col).value if col else None

            catalog.append({
                "sheet": sheet_name,
                "row": i,
                "item_code": cell("Item_Code") or "",
                "item_name": item_name,
                "item_description": cell("Item_Description") or "",
                "hsn": str(cell("HSN_Code") or "").strip(),
                "mrp": cell("MPR"),
                "quantity": cell("Quantity"),
                "gst_pct": cell("GST %"),
                "cost_price": cell("Cost_Price"),
                "sale_price": cell("Sale_PriceT1"),
            })
    return catalog


def build_full_description(sheet, item_name, item_description):
    text = f"{sheet} {item_name or ''} {item_description or ''}"
    return re.sub(r" +", " ", text).strip()


def load_customers(billdesk_path):
    wb = openpyxl.load_workbook(billdesk_path, data_only=True)
    sht = wb["CUSTPR"]
    headers = {}
    for j in range(1, sht.max_column + 1):
        h = sht.cell(row=1, column=j).value
        if h in CUSTOMER_COLUMNS:
            headers[h] = j

    customers = []
    for i in range(2, sht.max_row + 1):
        cid = sht.cell(row=i, column=headers.get("customerID", 3)).value
        if not cid:
            continue
        customers.append({
            "customer_id": cid,
            "name": sht.cell(row=i, column=headers.get("Customer_Name", 4)).value or "",
            "contact_number": sht.cell(row=i, column=headers.get("Contact_Number", 5)).value or "",
            "address_details": sht.cell(row=i, column=headers.get("Address_Details", 6)).value or "",
            "gstn": sht.cell(row=i, column=headers.get("GSTN", 7)).value or "",
        })
    return customers


# --------------------------------------------------------------------------- #
# Per-line rate/amount calculation (rateWithout_gst + adjust_purchaseList)
# --------------------------------------------------------------------------- #
def rate_without_gst(price, gst_pct):
    return round(float(price) * 100.0 / (100.0 + float(gst_pct)), 2)


def build_bill_lines(raw_lines, bill_type):
    """
    raw_lines: [{sheet, item_name, item_description, hsn, mrp, gst_pct,
                 sale_price, cost_price, qty}]  (qty = what's being sold now)
    Returns bill lines with rate/amount computed, plus a
    {description: cost_price} map for margin/CostReport purposes.
    """
    lines = []
    cost_map = {}
    for raw in raw_lines:
        description = build_full_description(raw["sheet"], raw["item_name"], raw.get("item_description"))
        gst_pct = float(raw["gst_pct"] or 0)
        sale_price = float(raw["sale_price"] or 0)
        qty = float(raw["qty"])
        taxable_rate = rate_without_gst(sale_price, gst_pct) if gst_pct or sale_price else 0.0

        if bill_type == "Tax Invoice":
            amount = round(taxable_rate * qty, 2)
        else:
            amount = round(sale_price * qty, 2)

        lines.append({
            "description": description,
            "hsn": str(raw.get("hsn") or "")[:4],
            "mrp": raw.get("mrp") or "",
            "qty": qty,
            "gst_pct": gst_pct,
            "rate": sale_price,          # GST-inclusive selling price/unit -> maps to 'r<L>'
            "taxable_rate": taxable_rate,  # excl-GST rate/unit             -> maps to 'w<L>'
            "amount": amount,             # line taxable amount            -> maps to 'a<L>'
            "sheet": raw["sheet"],
            "row": raw.get("row"),
            "product_id": raw.get("product_id") or raw.get("id"),
        })
        cost_map[description] = float(raw.get("cost_price") or 0)
    return lines, cost_map


# --------------------------------------------------------------------------- #
# Tax bucketing (get_tax_calculation + bottom_part, faithfully ported)
# --------------------------------------------------------------------------- #
def compute_tax(lines, bill_type):
    buckets = {}
    for line in lines:
        gst = int(line["gst_pct"])
        buckets.setdefault(gst, []).append(line["amount"])

    tax_fields = {}
    ttaxamt = cgst = sgst = 0.0
    for gst, amounts in buckets.items():
        if gst == 0:
            continue
        taxable = round(sum(amounts), 2)
        taxable_dec = Decimal(str(taxable))
        rate_dec = Decimal(str(gst)) / Decimal("100")
        half = (taxable_dec * rate_dec / Decimal("2")).quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)
        cg = float(half)
        sg = float(half)
        tot = round(cg + sg, 2)

        tax_fields[f"tax{gst}"] = taxable
        tax_fields[f"cg{gst}"] = cg
        tax_fields[f"sg{gst}"] = sg
        tax_fields[f"tot{gst}"] = tot

        ttaxamt += taxable
        cgst += cg
        sgst += sg

    # 0%-rated lines still count toward the taxable total, just contribute no tax.
    ttaxamt += round(sum(buckets.get(0, [])), 2)

    ttaxamt = round(ttaxamt, 2)
    cgst = round(cgst, 2)
    sgst = round(sgst, 2)
    total = round(ttaxamt + cgst + sgst, 2) if bill_type == "Tax Invoice" else ttaxamt

    return tax_fields, {
        "ttaxamt": ttaxamt, "cgst": cgst, "sgst": sgst,
        "total": total, "roff": "NA",
        "finalAmountWord": amount_to_words_inr(total),
    }


def amount_to_words_inr(amount):
    rupees = int(amount)
    paise = round((amount - rupees) * 100)
    words = num2words(rupees, lang="en_IN").capitalize()
    if paise > 0:
        paise_words = num2words(paise, lang="en_IN").lower()
        return f"{words} rupees {paise_words} paise only"
    return f"{words} rupees only"


def compute_hsn_summary(lines):
    """Groups line amounts by HSN code. Returns a list of {hsn, gst_pct, taxable_amt}."""
    grouped = {}
    for line in lines:
        key = line["hsn"]
        entry = grouped.setdefault(key, {"gst_pct": line["gst_pct"], "amounts": []})
        entry["amounts"].append(line["amount"])
    summary = []
    for hsn, entry in grouped.items():
        summary.append({
            "hsn": hsn,
            "gst_pct": entry["gst_pct"],
            "taxable_amt": round(sum(entry["amounts"]), 2),
        })
    return summary


def compute_margin(lines, cost_map):
    """qty * (rate - cost) per line - matches record_room_lookup.compute_margin_for_bill_json."""
    total = 0.0
    for line in lines:
        cost = cost_map.get(line["description"], 0.0)
        total += line["qty"] * (line["rate"] - cost)
    return round(total, 2)


# --------------------------------------------------------------------------- #
# Invoice numbering
# --------------------------------------------------------------------------- #
def next_invoice_number(sales_log_path, series_prefix_hint=None):
    """
    Ports get_indexNum()'s logic (front = all-but-last-4-chars, back =
    int(last4)+1, zero-padded) but sources the "previous" invoice number
    from the last row of sales_log.xlsx instead of indexBook/invoiceNum.txt,
    since that file wasn't provided. Drop a real invoiceNum.txt into
    indexBook/ to switch to exact-parity numbering later.
    """
    if not os.path.exists(sales_log_path):
        base = series_prefix_hint or "TS2026AA"
        return base + "0001"

    wb = openpyxl.load_workbook(sales_log_path, data_only=True)
    sht = wb["Sheet1"]
    last_invoice = None
    for row in range(sht.max_row, 1, -1):
        val = sht.cell(row=row, column=4).value  # column D = Invoice#
        if val and str(val).strip() and "RECORD DELEATED" not in str(sht.cell(row=row, column=2).value or ""):
            last_invoice = str(val).strip()
            break

    if not last_invoice or len(last_invoice) <= 4:
        base = series_prefix_hint or "TS2026AA"
        return base + "0001"

    front, back = last_invoice[:-4], last_invoice[-4:]
    try:
        back_num = int(back) + 1
    except ValueError:
        back_num = 1
    return front + str(back_num).zfill(4)


# --------------------------------------------------------------------------- #
# record_room JSON (lettered-field format, matches the real invoice sample)
# --------------------------------------------------------------------------- #
def build_final_data(customer, lines, invoice_no, date_str, bill_type):
    tax_fields, totals = compute_tax(lines, bill_type)
    hsn_summary = compute_hsn_summary(lines)

    final = {
        "customerGSTN": customer.get("gstn") or "Unregistered",
        "customerID": customer.get("customer_id") or "NA",
        "customerName": customer.get("name", ""),
        "customerContNum": customer.get("contact_number", ""),
        "CustomerDetails": customer.get("address_details", ""),
        "Date": date_str,
        "Invoice": invoice_no,
    }

    for i, line in enumerate(lines):
        letter = rrl.ITEM_LETTERS[i]
        final[f"item{letter}"] = line["description"]
        final[f"hs{letter}"] = line["hsn"]
        final[f"m{letter}"] = line["mrp"]
        final[f"q{letter}"] = line["qty"]
        final[f"g{letter}"] = line["gst_pct"]
        final[f"r{letter}"] = line["rate"]
        final[f"w{letter}"] = line["taxable_rate"]
        final[f"a{letter}"] = line["amount"]

    final.update(tax_fields)
    final.update(totals)

    final["hsn"] = "\n".join(str(h["hsn"]) for h in hsn_summary)
    final["gsh"] = "\n".join(str(h["gst_pct"]) for h in hsn_summary)
    final["hsnc_tax"] = "\n".join(str(h["taxable_amt"]) for h in hsn_summary)

    final["fileName"] = final["Invoice"] + "_" + final["Date"].replace("/", "")
    return final


def write_record_room_json(final_data):
    folder = rrl.local_folder_name(final_data["Date"])
    directory = os.path.join(RECORD_ROOM_BASE, "record_room", folder)
    os.makedirs(directory, exist_ok=True)
    path = os.path.join(directory, final_data["fileName"] + ".json")
    with open(path, "w") as f:
        json.dump(final_data, f)
    return path


def write_cost_report(final_data, cost_map):
    """{item_description: cost_price} - lets record_room_lookup recompute margin later."""
    folder = rrl.local_folder_name(final_data["Date"])
    directory = os.path.join(RECORD_ROOM_BASE, "record_room", folder)
    os.makedirs(directory, exist_ok=True)
    path = os.path.join(directory, final_data["fileName"] + "_CostReport.json")
    with open(path, "w") as f:
        json.dump(cost_map, f)
    return path


# --------------------------------------------------------------------------- #
# Stock decrement
# --------------------------------------------------------------------------- #
def decrement_stock(billdesk_path, sold_lines):
    """
    sold_lines: [{sheet, row, qty}] - subtracts qty from the "Quantity"
    (live stock) column for each row. Does not touch CurrentQTY (that field
    is the desktop tool's own cart-entry scratchpad; this web form doesn't
    use it). Negative resulting stock is allowed (your real data already
    has negative Quantity values in places), just returned as a warning.
    """
    wb = openpyxl.load_workbook(billdesk_path, read_only=False)
    warnings = []
    by_sheet = {}
    for line in sold_lines:
        by_sheet.setdefault(line["sheet"], []).append(line)

    for sheet_name, sheet_lines in by_sheet.items():
        sht = wb[sheet_name]
        qty_col = None
        for j in range(1, sht.max_column + 1):
            if sht.cell(row=1, column=j).value == "Quantity":
                qty_col = j
                break
        if not qty_col:
            continue
        for line in sheet_lines:
            current = sht.cell(row=line["row"], column=qty_col).value
            current = float(current) if current not in (None, "") else 0.0
            new_qty = current - float(line["qty"])
            if new_qty < 0:
                warnings.append(f"{sheet_name} row {line['row']}: stock now {new_qty:g} (oversold)")
            sht.cell(row=line["row"], column=qty_col).value = new_qty

    wb.save(billdesk_path)
    return warnings


# --------------------------------------------------------------------------- #
# sales_log.xlsx row (append for new, update-in-place for edit)
# --------------------------------------------------------------------------- #
def append_sales_log_row(sales_log_path, final_data, bill_type, margin_total):
    if not os.path.exists(sales_log_path):
        wb = openpyxl.Workbook()
        wb.active.title = "Sheet1"
        wb.save(sales_log_path)

    wb = openpyxl.load_workbook(sales_log_path, read_only=False)
    sht = wb["Sheet1"]
    row = sht.max_row + 1
    if sht.cell(row=1, column=1).value is None and row == 2:
        row = 2  # first real data row under an (assumed) blank/absent header

    _write_sales_log_row(sht, row, final_data, bill_type, margin_total)
    wb.save(sales_log_path)
    return row


def _write_sales_log_row(sht, row, final_data, bill_type, margin_total):
    sht.cell(row=row, column=1).value = final_data["fileName"]
    sht.cell(row=row, column=2).value = bill_type
    sht.cell(row=row, column=3).value = final_data["Date"]
    sht.cell(row=row, column=4).value = final_data["Invoice"]
    sht.cell(row=row, column=5).value = final_data["customerID"]
    sht.cell(row=row, column=6).value = final_data["customerName"]
    sht.cell(row=row, column=7).value = final_data["total"]
    sht.cell(row=row, column=8).value = margin_total
    sht.cell(row=row, column=9).value = "unpaid"
    sht.cell(row=row, column=10).value = None   # J: amount breakup - blank, not split
    sht.cell(row=row, column=11).value = final_data.get("ttaxamt")
    sht.cell(row=row, column=12).value = None   # L: taxable breakup - blank, not split
    sht.cell(row=row, column=13).value = None   # M: margin breakup - blank, not split
    sht.cell(row=row, column=15).value = "Candidate"  # O
