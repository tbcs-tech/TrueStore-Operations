"""
migration.py
=============
Imports the existing Excel-based data into the SQLite database, which
becomes the live store going forward (Excel remains useful for import/export
and as a human-readable backup, but is no longer read at request time).

Idempotent: products/customers/vendors are upserted by their natural key
(sheet+name / customer_id / vendor_id), and bills are matched by file_name,
so running this again after uploading an updated sheet won't create
duplicates - it'll update existing rows and add new ones.
"""
import openpyxl

import db
import billing_engine
from data_processor import load_invoices


def migrate_products_and_parties(billdesk_path):
    """Imports every product sheet + CUSTPR + VENDOR from billdesk.xlsx."""
    report = {"products": 0, "customers": 0, "vendors": 0, "errors": []}

    try:
        catalog = billing_engine.load_product_catalog(billdesk_path)
        for p in catalog:
            db.upsert_product(
                sheet=p["sheet"], item_code=p["item_code"], item_name=p["item_name"],
                item_description=p["item_description"], hsn=p["hsn"], mrp=p["mrp"],
                quantity=p["quantity"], gst_pct=p["gst_pct"],
                cost_price=p["cost_price"], sale_price=p["sale_price"],
            )
        report["products"] = len(catalog)
    except Exception as e:
        report["errors"].append(f"products: {e}")

    try:
        wb = openpyxl.load_workbook(billdesk_path, data_only=True)

        if "CUSTPR" in wb.sheetnames:
            sht = wb["CUSTPR"]
            headers = {}
            for j in range(1, sht.max_column + 1):
                h = sht.cell(row=1, column=j).value
                if h in ("Super Search", "customerID", "Customer_Name", "Contact_Number", "Address_Details", "GSTN"):
                    headers[h] = j
            count = 0
            for i in range(2, sht.max_row + 1):
                cid = sht.cell(row=i, column=headers.get("customerID", 3)).value
                if not cid:
                    continue
                db.upsert_customer(
                    customer_id=str(cid),
                    name=sht.cell(row=i, column=headers.get("Customer_Name", 4)).value or "",
                    contact_number=str(sht.cell(row=i, column=headers.get("Contact_Number", 5)).value or ""),
                    address_details=sht.cell(row=i, column=headers.get("Address_Details", 6)).value or "",
                    gstn=sht.cell(row=i, column=headers.get("GSTN", 7)).value or "",
                    category=str(sht.cell(row=i, column=headers.get("Super Search", 1)).value or "").strip().lower(),
                )
                count += 1
            report["customers"] = count
    except Exception as e:
        report["errors"].append(f"customers: {e}")

    try:
        wb = openpyxl.load_workbook(billdesk_path, data_only=True)
        if "VENDOR" in wb.sheetnames:
            sht = wb["VENDOR"]
            headers = {}
            for j in range(1, sht.max_column + 1):
                h = sht.cell(row=1, column=j).value
                if h in ("Super Search", "vendorID", "Vendor_Name", "Contact_Number", "Address_Details", "GSTN"):
                    headers[h] = j
            count = 0
            for i in range(2, sht.max_row + 1):
                vid = sht.cell(row=i, column=headers.get("vendorID", 3)).value
                if not vid:
                    continue
                db.upsert_vendor(
                    vendor_id=str(vid),
                    name=sht.cell(row=i, column=headers.get("Vendor_Name", 4)).value or "",
                    contact_number=str(sht.cell(row=i, column=headers.get("Contact_Number", 5)).value or ""),
                    address_details=sht.cell(row=i, column=headers.get("Address_Details", 6)).value or "",
                    gstn=sht.cell(row=i, column=headers.get("GSTN", 7)).value or "",
                    category=str(sht.cell(row=i, column=headers.get("Super Search", 1)).value or "").strip(),
                )
                count += 1
            report["vendors"] = count
    except Exception as e:
        report["errors"].append(f"vendors: {e}")

    return report


def migrate_bills(sales_log_path):
    """
    Imports every bill (paid and unpaid, split legs already exploded into
    their own rows by load_invoices) from sales_log.xlsx into the bills
    table. Skips file_names already present so this is safe to re-run.

    A handful of file_name collisions have been observed in real sales_log
    data (same invoice number + date appearing on more than one row - a
    pre-existing data-quality issue, not introduced by this migration).
    Since file_name must be unique (it's the real record_room join key),
    only the LAST occurrence of a colliding file_name is imported; the
    earlier one(s) are reported under "duplicates_skipped" so they can be
    reviewed in the source sheet if desired.
    """
    report = {"imported": 0, "skipped_existing": 0, "duplicates_skipped": [], "errors": []}

    try:
        raw_invoices = load_invoices(sales_log_path, include_all=True)
    except Exception as e:
        report["errors"].append(f"could not read sales_log: {e}")
        return report

    existing = {b["file_name"] for b in db.list_bills(include_deleted=True)}

    # Keep only the last occurrence of each file_name (dict preserves insertion
    # order in modern Python; re-assigning a key moves its value, not its position,
    # so we track order separately via a plain list built after dedup).
    by_file_name = {}
    for inv in raw_invoices:
        file_name = f"{inv['invoice_id']}_{inv['date_str'].replace('/', '')}" if inv["date_str"] else inv["invoice_id"]
        if file_name in by_file_name:
            report["duplicates_skipped"].append(file_name)
        by_file_name[file_name] = inv

    for file_name, inv in by_file_name.items():
        if file_name in existing:
            report["skipped_existing"] += 1
            continue
        try:
            split_leg = inv["invoice_id"][len(inv["original_invoice_id"]):] if inv["split"] else ""
            db.insert_bill(
                file_name=file_name,
                invoice_no=inv["invoice_id"],
                original_invoice_no=inv["original_invoice_id"],
                split_leg=split_leg,
                bill_type="Tax Invoice",
                date=inv["date_str"],
                customer_id=inv.get("customer_id", ""),
                customer_name=inv["customer"],
                total=inv["amount"],
                margin=inv["margin"],
                taxable_total=inv["taxable_amount"],
                status=inv["raw_status"],
                payment_date=None,
                is_candidate=inv["is_candidate_row"],
                created_by="migration",
                lines=[],  # line items live in record_room JSON already, not duplicated here
            )
            report["imported"] += 1
        except Exception as e:
            report["errors"].append(f"{inv['invoice_id']}: {e}")

    return report


def run_full_migration(billdesk_path, sales_log_path):
    return {
        "catalog": migrate_products_and_parties(billdesk_path),
        "bills": migrate_bills(sales_log_path),
    }
